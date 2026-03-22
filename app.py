import streamlit as st
import pandas as pd
import os
from datetime import datetime, timedelta
from io import BytesIO

# --- НАСТРОЙКИ ---
FILE_DB = 'database.csv'
STAFF_MEMBERS = ['Катя', 'Женя', 'Юля']
SERVICE_TYPES = ['Запись на студию', 'Урок по вокалу (Абонемент)', 'Пробный урок']
PACKET_TYPES = ['LITE', 'STANDARD', 'PRO', 'РАЗОВОЕ']

# --- ФУНКЦИИ ---

def load_data():
    """Безопасная загрузка данных"""
    if not os.path.exists(FILE_DB):
        # Возвращаем пустой DataFrame с правильными колонками, если файла нет
        return pd.DataFrame(columns=['date', 'time', 'staff', 'service', 'packet', 'client_name', 'phone', 'duration', 'status'])
    
    try:
        df = pd.read_csv(FILE_DB)
        return df
    except Exception as e:
        st.error(f"Ошибка чтения файла: {e}")
        return pd.DataFrame(columns=['date', 'time', 'staff', 'service', 'packet', 'client_name', 'phone', 'duration', 'status'])

def save_data(df):
    """Сохранение данных"""
    df.to_csv(FILE_DB, index=False)

def check_availability(df, date_str, start_time_str, duration_hours):
    """Проверка на конфликты"""
    if df.empty:
        return True, "OK"

    all_slots = [f"{h:02d}:{m:02d}" for h in range(9, 22) for m in [0, 30]]
    
    try:
        start_idx = all_slots.index(start_time_str)
    except ValueError:
        return False, "Неверное время начала"

    slots_needed = int(duration_hours * 2)
    required_slots = all_slots[start_idx : start_idx + slots_needed]

    if len(required_slots) < slots_needed:
        return False, "Запись выходит за пределы рабочего дня"

    day_records = df[df['date'] == date_str]
    if day_records.empty:
        return True, "OK"

    occupied_slots = []
    for _, row in day_records.iterrows():
        r_start = row['time']
        try:
            r_dur = int(float(row.get('duration', 1)) * 2)
            r_idx = all_slots.index(r_start)
            r_slots = all_slots[r_idx : r_idx + r_dur]
            occupied_slots.extend(r_slots)
        except (ValueError, TypeError):
            continue

    conflict_slots = set(required_slots) & set(occupied_slots)
    
    if conflict_slots:
        conflicting_record = day_records[day_records['time'].isin(list(conflict_slots))]
        who = conflicting_record.iloc[0]['staff'] if not conflicting_record.empty else "Неизвестно"
        client = conflicting_record.iloc[0]['client_name'] if not conflicting_record.empty else ""
        return False, f"Конфликт! Время занято: {who} (клиент {client})"
    
    return True, "OK"

def generate_excel(df):
    """Генерация Excel файла"""
    if df.empty:
        # Создаем пустой файл с заголовками, если данных нет
        df_empty = pd.DataFrame(columns=['date', 'time', 'staff', 'service', 'packet', 'client_name', 'phone', 'duration', 'status'])
        df_to_write = df_empty
    else:
        df_to_write = df.sort_values(by=['date', 'time', 'staff'])

    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_to_write.to_excel(writer, index=False, sheet_name='Расписание')
        
        workbook = writer.book
        worksheet = writer.sheets['Расписание']
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            
        # Автоширина
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Раскраска строк
        if not df_to_write.empty:
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                # Индекс сотрудника в нашем списке колонок: 2 (третья колонка)
                staff_cell = row[2] 
                staff_name = str(staff_cell.value) if staff_cell.value else ""
                
                fill_color = None
                text_color = "000000"
                
                if 'Катя' in staff_name:
                    fill_color = "3b82f6"; text_color = "FFFFFF"
                elif 'Женя' in staff_name:
                    fill_color = "22c55e"; text_color = "FFFFFF"
                elif 'Юля' in staff_name:
                    fill_color = "f97316"; text_color = "FFFFFF"
                elif 'Абонемент' in staff_name:
                    fill_color = "eab308"; text_color = "000000"
                
                if fill_color:
                    cell_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell_font = Font(color=text_color, bold=True)
                    for cell in row:
                        cell.fill = cell_fill
                        cell.font = cell_font
                        cell.alignment = center_align

        worksheet.freeze_panes = "A2"

    return output.getvalue()

# --- ИНТЕРФЕЙС ---

st.set_page_config(page_title="Студия: Бронирование", layout="wide")
st.title("📅 Расписание студии")

# Загружаем данные
df = load_data()

# Боковая панель
st.sidebar.header("➕ Новая запись")

with st.sidebar.form("booking_form"):
    sel_date = st.date_input("Дата", value=datetime.today())
    
    time_slots = [f"{h:02d}:{m:02d}" for h in range(9, 22) for m in [0, 30]]
    sel_time = st.selectbox("Время начала", time_slots)
    
    sel_duration = st.selectbox("Длительность (часы)", [0.5, 1.0, 1.5, 2.0], index=1)
    sel_staff = st.selectbox("Сотрудник", STAFF_MEMBERS)
    
    # Автовыбор услуги
    default_svc_idx = 0
    if sel_staff == 'Юля': default_svc_idx = 2
    elif 'Абонемент' in str(sel_staff): default_svc_idx = 1 # Упрощено
    
    sel_service = st.selectbox("Тип услуги", SERVICE_TYPES, index=default_svc_idx)
    sel_packet = st.selectbox("Пакет", PACKET_TYPES)
    
    client_name = st.text_input("Имя клиента")
    client_phone = st.text_input("Телефон")
    status = st.selectbox("Оплата", ["Предоплата+", "Ожидает", "Оплачено"])
    
    submitted = st.form_submit_button("💾 Сохранить")

    if submitted:
        if not client_name:
            st.error("Введите имя клиента!")
        else:
            date_str = str(sel_date)
            is_free, msg = check_availability(df, date_str, sel_time, sel_duration)
            
            if is_free:
                new_record = {
                    'date': date_str, 'time': sel_time, 'staff': sel_staff,
                    'service': sel_service, 'packet': sel_packet,
                    'client_name': client_name, 'phone': client_phone,
                    'duration': sel_duration, 'status': status
                }
                new_df = pd.DataFrame([new_record])
                df = pd.concat([df, new_df], ignore_index=True)
                save_data(df)
                st.success("Запись добавлена!")
                st.rerun()
            else:
                st.error(f"⛔ {msg}")

# Основная часть
st.subheader("📊 Текущие записи")

if df.empty:
    st.info("Записей пока нет. Добавьте первую через меню слева.")
else:
    # Фильтры
    c1, c2 = st.columns(2)
    with c1:
        filter_staff = st.multiselect("Сотрудник", STAFF_MEMBERS, default=STAFF_MEMBERS)
    with c2:
        search_text = st.text_input("Поиск по имени клиента")
    
    df_view = df[df['staff'].isin(filter_staff)]
    if search_text:
        df_view = df_view[df_view['client_name'].str.contains(search_text, case=False)]
    
    # Отображение с цветом
    def highlight_rows(row):
        staff = str(row['staff'])
        service = str(row['service'])
        color = '#ffffff'
        txt = '#000000'
        
        if 'Катя' in staff: color, txt = '#3b82f6', '#ffffff'
        elif 'Женя' in staff: color, txt = '#22c55e', '#ffffff'
        elif 'Юля' in staff: color, txt = '#f97316', '#ffffff'
        elif 'Абонемент' in service: color, txt = '#eab308', '#000000'
        
        return [f'background-color: {color}; color: {txt}; font-weight: bold;'] * len(row)

    st.dataframe(df_view.style.apply(highlight_rows, axis=1), use_container_width=True)

# Экспорт
st.divider()
col_btn, col_txt = st.columns([1, 3])
with col_btn:
    if not df.empty:
        excel_data = generate_excel(df)
        st.download_button(
            label="📥 Скачать Excel отчет",
            data=excel_data,
            file_name=f"Report_{datetime.today().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.button("📥 Скачать Excel (пустой шаблон)", disabled=True)
with col_txt:
    st.write("Нажмите кнопку, чтобы скачать оформленный файл для бухгалтерии.")

